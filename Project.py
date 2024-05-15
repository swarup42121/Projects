import chat
from tkinter import *

import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook


def search_value():
    file_path = "a.xlsx"  # Replace with the actual path to your Excel file

    search_user = entry.get()
    search_pass = entry1.get()


    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        # Search for the value in the Excel file

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == search_user:
                    pass





        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == search_pass:


                    messagebox.showinfo("Login Successfull..!")






    except Exception as e:
        messagebox.showerror("Error", str(e))


# Create the main window
window = tk.Tk()
window.geometry("330x200")
# Create an entry widget for the user to enter a value


user_name = Label(window,
				text = "Username").place(x = 40,
										y = 5)
entry = tk.Entry(window)
entry.pack()
user_password = Label(window,
					text = "Password").place(x = 40,
											y = 20)
entry1 = tk.Entry(window)
entry1.pack()

# Create a button to trigger the search
button = tk.Button(window, text="Login", command=search_value)
button.pack()

# Run the Tkinter event loop
window.mainloop()
